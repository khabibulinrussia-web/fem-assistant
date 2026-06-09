"""Excel export for ФЭМ-ассистент"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO

from .models import Assumptions
from .engine import run_calculation

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
LABEL_FONT = Font(bold=True, size=10)
NORMAL_FONT = Font(size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_row(ws, row, cols, bold=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = THIN_BORDER
        cell.font = LABEL_FONT if bold else NORMAL_FONT
        if c == 1:
            cell.alignment = Alignment(horizontal="left")
        else:
            cell.alignment = Alignment(horizontal="center")


def generate_excel(a: Assumptions, data: dict) -> bytes:
    wb = Workbook()

    # Sheet 1: Assumptions
    ws = wb.active
    ws.title = "Assumptions"
    ws.cell(1, 1, "Параметр").font = HEADER_FONT
    ws.cell(1, 2, "Значение").font = HEADER_FONT
    rows_a = [
        ("Целевая выручка в день", a.target_revenue_per_day),
        ("Средний чек", a.avg_check),
        ("Заказов в день", a.orders_per_day),
        ("Ramp-up мес 1", a.ramp_month_1),
        ("Ramp-up мес 2", a.ramp_month_2),
        ("Ramp-up мес 3+", a.ramp_month_3),
        ("Себестоимость заказа", a.cost_per_order),
        ("ФОТ", a.payroll),
        ("Эквайринг %", a.acquiring_pct),
        ("Аренда", a.rent),
        ("Маркетинг", a.marketing),
        ("Логистика на заказ", a.logistics_per_order),
        ("Прочие расходы", a.other_expenses),
        ("Регистрация ИП", a.reg_ip),
        ("Оборудование", a.equipment),
        ("Сайт", a.website),
        ("Резервный фонд", a.reserve_fund),
        ("Налоговая ставка", a.tax_rate),
        ("Дни запаса", a.inventory_days),
        ("Отсрочка поставщика", a.supplier_deferral),
        ("Отсрочка клиента", a.customer_deferral),
    ]
    for i, (label, val) in enumerate(rows_a, 2):
        ws.cell(i, 1, label).font = NORMAL_FONT
        ws.cell(i, 2, val).font = NORMAL_FONT
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15

    months = [f"Мес {i}" for i in range(1, 37)]

    def write_table(ws, title, rows_data, row_start=1):
        ws.cell(row_start, 1, title).font = Font(bold=True, size=12)
        for c_idx, label in enumerate(["Показатель"] + months, 1):
            cell = ws.cell(row_start + 1, c_idx, label)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = THIN_BORDER
        for r_idx, (label, values) in enumerate(rows_data, row_start + 2):
            ws.cell(r_idx, 1, label).font = LABEL_FONT
            ws.cell(r_idx, 1).border = THIN_BORDER
            for c_idx, val in enumerate(values, 2):
                cell = ws.cell(r_idx, c_idx, round(val, 2))
                cell.font = NORMAL_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")
        return row_start + 2 + len(rows_data) + 1

    # Sheet 2: P&L
    ws2 = wb.create_sheet("P&L")
    pnl = data["pnl"]
    pnl_data = [
        ("Выручка", pnl["revenue"]),
        ("Эквайринг", pnl["acquiring"]),
        ("Чистая выручка", pnl["net_revenue"]),
        ("Себестоимость", pnl["cost_of_goods"]),
        ("ФОТ", pnl["payroll"]),
        ("Логистика", pnl["logistics"]),
        ("Маркетинг", pnl["marketing"]),
        ("Аренда", pnl["rent"]),
        ("Прочие", pnl["other"]),
        ("Итого расходов", pnl["total_expenses"]),
        ("Прибыль до налога", pnl["profit_before_tax"]),
        ("Налог", pnl["tax"]),
        ("Чистая прибыль", pnl["net_profit"]),
        ("Накопленная прибыль", pnl["cumulative_profit"]),
    ]
    write_table(ws2, "Отчёт о прибылях и убытках", pnl_data)

    # Sheet 3: Cash Flow
    ws3 = wb.create_sheet("Cash Flow")
    cf = data["cashflow"]
    cf_data = [
        ("Поступления", cf["operating_inflow"]),
        ("Платежи (-)", cf["operating_outflow"]),
        ("Чистый опер. поток", cf["net_operating"]),
        ("Инвестиции", cf["investing"]),
        ("Финансирование", cf["financing"]),
        ("ДС на начало", cf["cash_start"]),
        ("ДС на конец", cf["cash_end"]),
        ("ЧДПС", cf["net_cashflow"]),
    ]
    write_table(ws3, "Бюджет движения денежных средств", cf_data)

    # Sheet 4: Balance Sheet
    ws4 = wb.create_sheet("Balance Sheet")
    bs = data["balance_sheet"]
    bs_data = [
        ("Денежные средства", bs["cash"]),
        ("Запасы", bs["inventory"]),
        ("Основные средства", bs["fixed_assets"]),
        ("Итого активы", bs["total_assets"]),
        ("Капитал", bs["capital"]),
        ("Кредиторская задолж.", bs["accounts_payable"]),
        ("Итого пассивы", bs["total_liabilities"]),
    ]
    write_table(ws4, "Баланс", bs_data)

    # Sheet 5: Ratios
    ws5 = wb.create_sheet("Ratios")
    r = data["ratios"]
    ws5.cell(1, 1, "Показатель").font = HEADER_FONT
    ws5.cell(1, 2, "Значение").font = HEADER_FONT
    ws5.cell(1, 3, "Норма").font = HEADER_FONT
    _style_header(ws5, 1, 3)
    ratio_rows = [
        ("ROE (рентабельность капитала)", r["ROE"], "> 15%"),
        ("ROCE (рентабельность капитала)", r["ROCE"], "> 12%"),
        ("OPM (операционная маржа)", r["OPM"], "> 10%"),
        ("GPM (валовая маржа)", r["GPM"], "> 20%"),
        ("Проверка баланса", r["BalanceCheck"], "0"),
    ]
    for i, (label, val, norm) in enumerate(ratio_rows, 2):
        ws5.cell(i, 1, label).font = NORMAL_FONT
        ws5.cell(i, 2, val).font = NORMAL_FONT
        ws5.cell(i, 3, norm).font = NORMAL_FONT
    ws5.column_dimensions["A"].width = 35
    ws5.column_dimensions["B"].width = 15
    ws5.column_dimensions["C"].width = 10

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
